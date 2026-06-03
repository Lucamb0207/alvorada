"""
Sincroniza producao_import.xlsx com o banco PostgreSQL do Render.

Uso:
    $env:DATABASE_URL = "postgresql://..."
    python sync_planilha.py

Salve o xlsx antes de rodar para garantir que as fórmulas estejam calculadas.
"""

import os
import sys
from datetime import datetime
import openpyxl
import db

XLSX_PATH = os.path.join(os.path.dirname(__file__), "producao_import.xlsx")

# Linha 1 = cabeçalhos internos, Linha 2 = labels legíveis, dados a partir da linha 3
HEADER_ROW = 1
DATA_START  = 3


def _int_or_none(val):
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(round(float(str(val).replace(",", "."))))
    except (ValueError, TypeError):
        return None


def load_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[HEADER_ROW]]

    def col(row_values, name):
        try:
            return row_values[headers.index(name)]
        except ValueError:
            return None

    rows = []
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        fecha_val = col(row, "fecha")
        if fecha_val is None:
            continue

        if isinstance(fecha_val, datetime):
            fecha = fecha_val.date()
        else:
            try:
                fecha = datetime.strptime(str(fecha_val).strip(), "%d-%m-%Y").date()
            except ValueError:
                print(f"  [aviso] Data inválida ignorada: {fecha_val}")
                continue

        rows.append({
            "fecha":            fecha,
            "pn_bls":           _int_or_none(col(row, "pn_bls")),
            "prom_mes_operada": _int_or_none(col(row, "prom_mes_operada")),
            "pdt_plan":         _int_or_none(col(row, "pdt_plan")),
            "var_vs_pdt":       _int_or_none(col(row, "var_vs_pdt")),
            "falhas":           str(col(row, "falhas") or "").strip(),
        })

    return rows


def main():
    print("=" * 55)
    print("  Sync Planilha -> Render DB")
    print("=" * 55)

    if not os.environ.get("DATABASE_URL"):
        print("\n❌ DATABASE_URL não definida. Execute primeiro:")
        print('   $env:DATABASE_URL = "postgresql://..."')
        sys.exit(1)

    print(f"\nLendo {XLSX_PATH} ...")
    rows = load_xlsx()
    print(f"  {len(rows)} linha(s) encontrada(s) na planilha.")

    if not rows:
        print("Nenhuma linha para importar.")
        sys.exit(0)

    print("\nPrévia das linhas a sincronizar:")
    for r in rows:
        print(f"  {r['fecha']}  PN={r['pn_bls']}  Prom={r['prom_mes_operada']}  PDT={r['pdt_plan']}  Var={r['var_vs_pdt']}")

    if "--sim" not in sys.argv:
        confirm = input("\nConfirmar sincronizacao? (s/n): ").strip().lower()
        if confirm != "s":
            print("Cancelado.")
            sys.exit(0)

    saved, errors = 0, []
    for r in rows:
        try:
            db.upsert_producao(r)
            saved += 1
        except Exception as exc:
            errors.append(f"  {r['fecha']}: {exc}")

    print(f"\n✅ {saved} linha(s) sincronizada(s) com sucesso.")
    if errors:
        print("⚠️  Erros:")
        for e in errors:
            print(e)


if __name__ == "__main__":
    main()
